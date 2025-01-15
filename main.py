import streamlit as st
import pandas as pd
import os
import matplotlib.pyplot as plt
from openpyxl import Workbook, load_workbook
import sys
sys.path.append('C:\\Users\\elbak\\Downloads\\tournament-management\\gender-distrub')
sys.path.append('C:\\Users\\elbak\\Downloads\\tournament-management\\distrub-club-score')
sys.path.append('C:\\Users\\elbak\\Downloads\\tournament-management\\genderscore')
sys.path.append('C:\\Users\\elbak\\Downloads\\tournament-management\\medals-club')
sys.path.append('C:\\Users\\elbak\\Downloads\\tournament-management\\techniaue')
sys.path.append('C:\\Users\\elbak\\Downloads\\tournament-management\\weight')
sys.path.append('C:\\Users\\elbak\\Downloads\\tournament-management\\distrubution-part-club')
from gender import gender_fighter
from scoreclub import score_destrub
from gensc import gender_score
from clubs_medals import clubs_medals
from tech import technique
from weight_class import weight_class
from distrib_partclub import club_part


def load_css(file_path):
    with open(file_path, "r") as f:
        css = f.read()
    st.markdown(f"<style>{css}</style>", unsafe_allow_html=True)

css_file = "C:\\Users\\elbak\\Downloads\\tournament-management\\style.css" 
if os.path.exists(css_file):
    load_css(css_file)
else:
    st.warning(f"CSS file '{css_file}' not found!")
EXCEL_FILE = "tournament_management.xlsx"
FIGHTERif_SHEET = "fighters-info"
MATCHES_SHEET = "matches"
fighterif_columns = [
    "fighter_name", "club", "gender", "cathegory_age", "poids_category"
    
]

matches_columns = [
    "fighter_name","match_number", "club" ,"gender", "cathegory_age", "poids_category", "score", "time", "hai_diem", "ba_diem", "mot_diem", "status"
]
if not os.path.exists(EXCEL_FILE):
    wb = Workbook()
    wb.create_sheet(FIGHTERif_SHEET, 0)
    wb.create_sheet(MATCHES_SHEET, 1)
    fighters_df = pd.DataFrame(columns=fighterif_columns)
    matches_df = pd.DataFrame(columns=matches_columns)
    fighters_df.to_excel(EXCEL_FILE, sheet_name=FIGHTERif_SHEET, index=False)
    matches_df.to_excel(EXCEL_FILE, sheet_name=MATCHES_SHEET, index=False)
    wb.save(EXCEL_FILE)
else:
    wb = load_workbook(EXCEL_FILE)
    fighters_df = pd.read_excel(EXCEL_FILE, sheet_name=FIGHTERif_SHEET)
    matches_df = pd.read_excel(EXCEL_FILE, sheet_name=MATCHES_SHEET)

def save_data():
    with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        fighters_df.to_excel(writer, sheet_name=FIGHTERif_SHEET, index=False)

def save_data_match():
    with pd.ExcelWriter(EXCEL_FILE, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
        matches_df.to_excel(writer, sheet_name=MATCHES_SHEET, index=False)
        
def graphics_analysis():
    st.sidebar.title("Menu Graphiques")

    if st.sidebar.button("Retour à l'accueil"):
        st.session_state["page"] = "accueil"
        st.session_state["page2"] = "graphics_analysis"
        st.session_state["page3"] = "graphics_analysis"
        st.session_state["page4"] = "graphics_analysis"
        st.session_state["page5"] = "graphics_analysis"
        st.session_state["page6"] = "graphics_analysis"
        st.session_state["page7"] = "graphics_analysis"
        st.session_state["page8"] = "graphics_analysis"
        st.rerun()

    page = st.sidebar.radio("Options", [
        "distrubution par sexe", "distrubution par club", "distrubution de score par sexes", 
        "distrubution de medailles par club", "distrubution de technique", "distrubution de poids","distrubution participant par club"
    ])

    if page == "distrubution par sexe":
        gender_fighter()
    elif page == "distrubution par club":
        score_destrub()
    elif page == "distrubution de score par sexes":
        gender_score()
    elif page == "distrubution de medailles par club":
        clubs_medals()
    elif page == "distrubution de technique":
        technique()
    elif page == "distrubution de poids":
        weight_class()
    elif page == "distrubution participant par club":
        club_part()
        
        
def accueil():
    st.title("Accueil")
    st.image(r"pic\\ahmeddd.jpg", use_container_width =True)
    st.write("Bienvenue sur StremXFighter, l'application qui facilite la gestion des combattants, l'organisation des matchs et l'analyse des données avec des graphiques intuitifs.")
    
    col1, col2 = st.columns(2)
    with col1:
        if st.button("Gestion des combattants"):
            st.session_state["page"] = "fighter_management"
            st.rerun()

    with col2:
        if st.button("Graphiques et Analyse"):
            st.session_state["page"] = "graphics_analysis"
            st.rerun()

def fighter_management():
    st.sidebar.title("Menu Combattants")
    if st.sidebar.button("Retour à l'accueil"):
        st.session_state["page"] = "accueil"
        st.rerun()


    page = st.sidebar.radio("Options", [
        "Ajouter un combattant", "Liste des combattants", 
        "Ajouter un match", "Liste des matchs", 
        "Mettre à jour le statut", "Supprimer un combattant"
    ])
    
    if page == "Ajouter un combattant":
        add_fighter()
    elif page == "Liste des combattants":
        list_fighters()
    elif page == "Ajouter un match":
        match_fighter()
    elif page == "Liste des matchs":
        list_match()
    elif page == "Mettre à jour le statut":
        fighter_status()
    elif page == "Supprimer un combattant":
        delete_fighter()


def add_fighter():
    st.header("Ajouter un combattant")
    fighter_name = st.text_input("Nom et prénom du combattant")
    club = st.text_input("Club du combattant")
    gender = st.selectbox("Genre", options=["Homme", "Femme"])
    category_age = st.selectbox("Catégorie d'âge", options=["Mini poussins", "Poussins", "Pupilles", "Benjamins", "Minimes", "Cadets", "Juniors"])
    
    poids_options = {
        "Mini poussins": ["-20kg", "-25kg", "-30kg", "-35kg", "+35kg"],
        "Poussins": ["-20kg", "-25kg", "-30kg", "-35kg", "-40kg", "+40kg"],
        "Pupilles": ["-25kg", "-30kg", "-35kg", "-40kg", "-45kg", "+45kg"],
        "Benjamins": ["-30kg", "-35kg", "-40kg", "-45kg", "-50kg", "+50kg"],
        "Minimes": ["-35kg", "-40kg", "-45kg", "-50kg", "-55kg", "+55kg"],
        "Cadets": ["-47kg", "-54kg", "+54kg"],
        "Juniors": ["-48kg", "-53kg", "-59kg", "+59kg"],
    }
    poids_category = st.selectbox("Catégorie de poids", options=poids_options[category_age])
    
    if st.button("Ajouter le combattant"):
        new_row = {
            "fighter_name": fighter_name,
            "club": club,
            "gender": gender,
            "cathegory_age": category_age,
            "poids_category": poids_category,

        }
        global fighters_df
        fighters_df = pd.concat([fighters_df, pd.DataFrame([new_row])], ignore_index=True)
        save_data()
        st.success(f"Combattant {fighter_name} ajouté avec succès.")
def list_fighters():
    st.header("Liste des combattants")
    if fighters_df.empty:
        st.warning("Aucun combattant ajouté pour le moment.")
    else:
        st.dataframe(fighters_df)
        
def match_fighter():
    global matches_df
    st.header("Ajouter un match")
    if fighters_df.empty:
        st.warning("Ajoutez des combattants avant d'enregistrer un match.")
        return

    fighter_name = st.selectbox("Sélectionnez un combattant", fighters_df["fighter_name"].tolist())
    score = st.number_input("Score", min_value=0)
    fighter_info = fighters_df[fighters_df["fighter_name"] == fighter_name].iloc[0]
    club = fighter_info["club"]
    gender=fighter_info["gender"]

    category_age = fighter_info["cathegory_age"]
    poids_category=fighter_info["poids_category"]

    time = st.number_input("Temps joué (en minutes)")
    hai_diem = st.number_input("Hai Diem", min_value=0)
    ba_diem = st.number_input("Ba Diem", min_value=0)
    mot_diem = st.number_input("Mot Diem", min_value=0)
    status = st.selectbox("Statut", options=["win", "lose"])
    
    if st.button("Ajouter le match"):
        new_match = {
            "fighter_name": fighter_name,
            "score": score,
            "time": time,
            "hai_diem": hai_diem,
            "ba_diem": ba_diem,
            "mot_diem": mot_diem,
            "status": status,
            "gender": gender,

            "club": club,
            "cathegory_age": category_age ,
            "poids_category": poids_category,

        }
        matches_df = pd.concat([matches_df, pd.DataFrame([new_match])], ignore_index=True)
        save_data_match()
        st.success(f"Match pour {fighter_name} ajouté avec succès.")

def list_match():
    st.header("Liste des matchs")
    if matches_df.empty:
        st.warning("Aucun match enregistré pour le moment.")
    else:
        st.dataframe(matches_df)

def fighter_status():
    st.header("Mettre à jour le statut du combattant")
    if fighters_df.empty:
        st.warning("Aucun combattant disponible.")
        return

    fighter_name = st.selectbox("Sélectionnez un combattant", fighters_df["fighter_name"].tolist())
    medal = st.selectbox("Médaille", options=["gold", "silver", "bronze", "None"])
    
    if st.button("Mettre à jour"):
        fighter_index = fighters_df[fighters_df["fighter_name"] == fighter_name].index[0]
        fighters_df.at[fighter_index, "medal"] = medal
        save_data()
        st.success(f"Statut du combattant {fighter_name} mis à jour avec succès.")
        
        
        
def delete_fighter():
    global fighters_df

    st.header("Supprimer un combattant")
    if fighters_df.empty:
        st.warning("Aucun combattant disponible pour suppression.")
        return

    fighter_name = st.selectbox("Sélectionnez un combattant à supprimer", fighters_df["fighter_name"].tolist())
    if st.button("Supprimer"):
        fighters_df = fighters_df[fighters_df["fighter_name"] != fighter_name]
        save_data()
        st.success(f"Combattant {fighter_name} supprimé avec succès !")



if "page" not in st.session_state:
    st.session_state["page"] = "accueil"

if st.session_state["page"] == "accueil":
    accueil()
elif st.session_state["page"] == "fighter_management":
    fighter_management()
elif st.session_state["page"] == "graphics_analysis":
    graphics_analysis()
